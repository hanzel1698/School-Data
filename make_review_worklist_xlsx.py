#!/usr/bin/env python3
"""
Convert schools_review_worklist.csv into schools_review_worklist.xlsx with
a real clickable hyperlink in the Maps Link column (plain CSV text URLs
aren't clickable in most spreadsheet apps unless retyped by hand).

Usage:
    python make_review_worklist_xlsx.py
"""

import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SOURCE_CSV = "schools_review_worklist.csv"
OUTPUT_XLSX = "schools_review_worklist.xlsx"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
BODY_FONT = Font(name=FONT_NAME)
LINK_FONT = Font(name=FONT_NAME, color="0563C1", underline="single")
FILLIN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

COLUMN_WIDTHS = {
    "School Code": 12,
    "School Name": 32,
    "Educational Sub-District": 20,
    "Local Body": 18,
    "Block Name": 16,
    "Latitude": 12,
    "Longitude": 12,
    "Maps Link": 14,
    "Verified (Y/N)": 14,
    "Corrected Latitude": 16,
    "Corrected Longitude": 17,
    "Notes": 30,
}
FILLIN_COLUMNS = {"Verified (Y/N)", "Corrected Latitude", "Corrected Longitude", "Notes"}


def main():
    with open(SOURCE_CSV, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        rows = list(reader)

    wb = Workbook()
    ws = wb.active
    ws.title = "Schools Review"

    for col_idx, name in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(name, 16)
    ws.freeze_panes = "A2"

    maps_col = fieldnames.index("Maps Link") + 1

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, name in enumerate(fieldnames, start=1):
            value = row[name]
            if name == "Maps Link":
                cell = ws.cell(row=row_idx, column=col_idx, value="Open in Maps")
                cell.hyperlink = value
                cell.font = LINK_FONT
                continue
            if name in ("Latitude", "Longitude", "Corrected Latitude", "Corrected Longitude") and value:
                cell = ws.cell(row=row_idx, column=col_idx, value=float(value))
            elif name == "School Code" and value:
                cell = ws.cell(row=row_idx, column=col_idx, value=int(value))
            else:
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            if name in FILLIN_COLUMNS:
                cell.fill = FILLIN_FILL

    ws.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}{len(rows) + 1}"

    wb.save(OUTPUT_XLSX)
    print(f"Wrote {len(rows)} rows to {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
