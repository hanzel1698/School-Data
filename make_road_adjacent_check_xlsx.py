#!/usr/bin/env python3
"""
Turn schools_road_adjacent_maps_check.csv into a color-coded, bucketed
.xlsx. Buckets (per the analysis discussed with the user):

  - Credible Error Candidate (300m-3km): confident name+management match,
    meaningfully far from the DB coordinate - the real actionable list.
  - Probably Match Noise (>3km): confident match but implausibly far for a
    real geocoding slip - more likely the fuzzy matcher hit a different,
    similarly-named real place. Deprioritized, not "fine".
  - Likely Correct (<=300m): confidently matched and close - low priority.
  - Low-Confidence / Unverifiable: Google has no distinguishing listing, or
    matched a different Govt/Aided institution - this method has no signal
    either way for these.

Usage:
    python make_road_adjacent_check_xlsx.py
"""

import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SOURCE_CSV = "schools_road_adjacent_maps_check.csv"
OUTPUT_XLSX = "schools_road_adjacent_maps_check.xlsx"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
BODY_FONT = Font(name=FONT_NAME)
LINK_FONT = Font(name=FONT_NAME, color="0563C1", underline="single")

BUCKET_ORDER = [
    "Credible Error Candidate (300m-3km)",
    "Probably Match Noise (>3km)",
    "Low-Confidence / Unverifiable",
    "Likely Correct (<=300m)",
]
BUCKET_FILL = {
    "Credible Error Candidate (300m-3km)": PatternFill("solid", fgColor="F8CBAD"),  # orange - check this
    "Probably Match Noise (>3km)": PatternFill("solid", fgColor="FFF2CC"),          # yellow - skip, likely noise
    "Low-Confidence / Unverifiable": PatternFill("solid", fgColor="D9D9D9"),        # gray - unknown
    "Likely Correct (<=300m)": PatternFill("solid", fgColor="C6E0B4"),              # green - fine
}


def bucket_for(row):
    if row["Notes"].startswith("LOW_CONFIDENCE_MATCH"):
        return "Low-Confidence / Unverifiable"
    dist = float(row["Distance Meters"])
    if dist <= 300:
        return "Likely Correct (<=300m)"
    if dist <= 3000:
        return "Credible Error Candidate (300m-3km)"
    return "Probably Match Noise (>3km)"


def main():
    with open(SOURCE_CSV, encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    for r in rows:
        r["Bucket"] = bucket_for(r)

    bucket_rank = {name: i for i, name in enumerate(BUCKET_ORDER)}
    rows.sort(key=lambda r: (
        bucket_rank[r["Bucket"]],
        -float(r["Distance Meters"]) if r["Distance Meters"] else 0,
    ))

    fieldnames = [
        "Bucket", "School Name", "Educational Sub-District", "Block Name",
        "DB Latitude", "DB Longitude", "DB Location",
        "Google Matched Name", "Google Latitude", "Google Longitude", "Google Location",
        "Distance Meters", "Name Similarity", "Management Mismatch", "Notes",
    ]
    col_widths = {
        "Bucket": 30, "School Name": 30, "Educational Sub-District": 20, "Block Name": 16,
        "DB Latitude": 12, "DB Longitude": 12, "DB Location": 12,
        "Google Matched Name": 32, "Google Latitude": 12, "Google Longitude": 12, "Google Location": 12,
        "Distance Meters": 14, "Name Similarity": 14, "Management Mismatch": 16, "Notes": 30,
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Road-Adjacent Schools Check"

    for col_idx, name in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(name, 16)
    ws.freeze_panes = "A2"

    for row_idx, r in enumerate(rows, start=2):
        fill = BUCKET_FILL[r["Bucket"]]
        for col_idx, name in enumerate(fieldnames, start=1):
            if name == "DB Location":
                cell = ws.cell(row=row_idx, column=col_idx, value="Open")
                cell.hyperlink = f"http://maps.google.com/maps?q={r['DB Latitude']},{r['DB Longitude']}"
                cell.font = LINK_FONT
            elif name == "Google Location":
                if r["Google Latitude"] and r["Google Longitude"]:
                    cell = ws.cell(row=row_idx, column=col_idx, value="Open")
                    cell.hyperlink = f"http://maps.google.com/maps?q={r['Google Latitude']},{r['Google Longitude']}"
                    cell.font = LINK_FONT
                else:
                    cell = ws.cell(row=row_idx, column=col_idx, value="")
                    cell.font = BODY_FONT
            else:
                value = r.get(name, "")
                if name in ("DB Latitude", "DB Longitude", "Google Latitude", "Google Longitude",
                            "Distance Meters", "Name Similarity") and value:
                    value = float(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = BODY_FONT
            cell.fill = fill

    ws.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}{len(rows) + 1}"

    # Legend sheet
    legend = wb.create_sheet("Legend")
    legend.column_dimensions["A"].width = 34
    legend.column_dimensions["B"].width = 70
    legend_rows = [
        ("Bucket", "Meaning"),
        ("Credible Error Candidate (300m-3km)", "Confident name+management match, meaningfully far from the DB coordinate. The actionable list - check these in Google Earth first."),
        ("Probably Match Noise (>3km)", "Confident match but implausibly far for a real geocoding slip - more likely the matcher found a different, similarly-named real place. Deprioritized."),
        ("Low-Confidence / Unverifiable", "Google has no distinguishing listing for this school, or matched a different Govt/Aided institution. This check has no signal either way."),
        ("Likely Correct (<=300m)", "Confidently matched and close to the DB coordinate. Low priority for review."),
    ]
    for r_idx, (bucket, meaning) in enumerate(legend_rows, start=1):
        b = legend.cell(row=r_idx, column=1, value=bucket)
        m = legend.cell(row=r_idx, column=2, value=meaning)
        if r_idx == 1:
            b.font = HEADER_FONT
            m.font = HEADER_FONT
            b.fill = HEADER_FILL
            m.fill = HEADER_FILL
        else:
            b.font = BODY_FONT
            m.font = BODY_FONT
            b.fill = BUCKET_FILL[bucket]
            m.fill = BUCKET_FILL[bucket]
        m.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUTPUT_XLSX)

    from collections import Counter
    counts = Counter(r["Bucket"] for r in rows)
    print(f"Wrote {len(rows)} rows to {OUTPUT_XLSX}")
    for name in BUCKET_ORDER:
        print(f"  {name}: {counts[name]}")


if __name__ == "__main__":
    main()
