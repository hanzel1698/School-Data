#!/usr/bin/env python3
"""
One-off transform of the in-progress schools_review_worklist.xlsx:
merge the separate "Corrected Latitude" / "Corrected Longitude" columns
into a single "Corrected Lat, Long" column, so a "lat, long" pair copied
straight out of Google Maps can be pasted into one cell. Preserves every
value the user has already entered (Verified, Corrected*, Notes).

Usage:
    python merge_corrected_latlon_columns.py
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

PATH = "schools_review_worklist.xlsx"
FILLIN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def combine(lat_val, lon_val):
    if lat_val is None and lon_val is None:
        return None
    if lat_val is not None and lon_val is not None:
        # Already a combined "lat, long" pasted into the lat cell by mistake/habit.
        if isinstance(lat_val, str) and "," in lat_val:
            return lat_val
        return f"{lat_val}, {lon_val}"
    return lat_val if lat_val is not None else lon_val


def main():
    wb = load_workbook(PATH)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    lat_col = headers.index("Corrected Latitude") + 1
    lon_col = headers.index("Corrected Longitude") + 1

    merged_values = {}
    for r in range(2, ws.max_row + 1):
        lat_val = ws.cell(row=r, column=lat_col).value
        lon_val = ws.cell(row=r, column=lon_col).value
        merged_values[r] = combine(lat_val, lon_val)

    # Delete the longitude column first (higher index) so the latitude
    # column's index stays valid; it becomes the merged column.
    ws.delete_cols(lon_col, 1)

    ws.cell(row=1, column=lat_col, value="Corrected Lat, Long")
    for r, val in merged_values.items():
        cell = ws.cell(row=r, column=lat_col, value=val)
        cell.fill = FILLIN_FILL

    ws.column_dimensions[get_column_letter(lat_col)].width = 30

    wb.save(PATH)
    print(f"Merged into single 'Corrected Lat, Long' column, preserved {sum(1 for v in merged_values.values() if v)} existing values.")


if __name__ == "__main__":
    main()
